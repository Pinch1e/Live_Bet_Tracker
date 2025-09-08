import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('betting_summary.xlsx')
ws = wb.active

# Add new column headers
ws['I1'] = 'Losses'
ws['J1'] = 'Odds'
ws['K1'] = 'How much placed'
ws['L1'] = 'Cash out'

# For the existing row (assuming row 2), add sample data
ws['I2'] = 0  # Losses
ws['J2'] = 2.5  # Odds
ws['K2'] = 10000  # How much placed
ws['L2'] = 10000  # Cash out

# Save the workbook
wb.save('betting_summary.xlsx')

print("Excel file updated with new columns.")
